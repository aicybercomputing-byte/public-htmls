#!/usr/bin/env python3
"""
sync-box-csv.py
Download one Box-hosted .xlsx (see XLSX_URL) and export each of its sheets
to the matching CSV in the repo (see SHEET_EXPORTS). "ai-x-events" and
"community-events" are tabs in the same workbook, not separate files.

The workbook's Box shared link is public with downloads enabled, so this
just does a plain HTTP GET on its "Direct link" (Share -> Link Settings ->
Direct link) — no Box app, OAuth, or credentials needed. If that link is
ever locked down, get a new "Direct link" from Box and update XLSX_URL.

The existing hourly push script handles git commit/push — this script only
needs to get the file on disk.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SETUP (one time)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Install dependencies:
     pip install -r requirements.txt

2. Run it:
     python3 scripts/sync-box-csv.py

3. Add to crontab (or call from your existing push script):
     */30 * * * * cd /path/to/public-htmls && python3 scripts/sync-box-csv.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import csv
import io
import pathlib
import sys
import urllib.error
import urllib.request

from openpyxl import load_workbook

REPO_ROOT = pathlib.Path(__file__).parent.parent

# ── Edit these to match your setup ───────────────────────────────────────
# Box "Direct link" for the workbook (Share -> Link Settings -> Direct link).
XLSX_URL = "https://usf.box.com/shared/static/icsswsoc5999tfs47kqqhbs5wdd7hi5w.xlsx"

# Sheet name (exact, case-sensitive) -> destination CSV path.
SHEET_EXPORTS = {
    "ai-x-events": REPO_ROOT / "ai-x" / "ai-x-events.csv",
    "community-events": REPO_ROOT / "ai-x" / "community-events.csv",
}
# ─────────────────────────────────────────────────────────────────────────


def download_workbook(url):
    try:
        with urllib.request.urlopen(url, timeout=30) as r:
            return r.read()
    except urllib.error.URLError as e:
        raise RuntimeError(f"Download failed: {e}") from e


def cell_to_str(v):
    if v is None:
        return ""
    # openpyxl reads whole numbers (e.g. a "day" column) as float (16.0);
    # avoid writing "16.0" into the CSV.
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def export_sheet_to_csv(ws, dest):
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue
            writer.writerow([cell_to_str(v) for v in row])
    print(f"Saved  -> {dest}")


def main():
    try:
        content = download_workbook(XLSX_URL)
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except RuntimeError as e:
        sys.exit(f"Error downloading workbook: {e}")

    failed = []
    for sheet_name, dest in SHEET_EXPORTS.items():
        if sheet_name not in wb.sheetnames:
            print(
                f"Error exporting {dest.name}: sheet '{sheet_name}' not found "
                f"(workbook has: {', '.join(wb.sheetnames)})",
                file=sys.stderr,
            )
            failed.append(dest.name)
            continue
        try:
            export_sheet_to_csv(wb[sheet_name], dest)
        except OSError as e:
            print(f"Error writing {dest}: {e}", file=sys.stderr)
            failed.append(dest.name)

    if failed:
        sys.exit(f"Failed to sync: {', '.join(failed)}")


if __name__ == "__main__":
    main()
