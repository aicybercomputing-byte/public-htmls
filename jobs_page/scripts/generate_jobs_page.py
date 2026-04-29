from __future__ import annotations

import argparse
import html
import json
import math
import re
import sys
from itertools import zip_longest
from pathlib import Path
from typing import Dict, List, Tuple

CONTENT_WORKBOOK_NAME = "jobs_page_content.xlsx"


REQUIRED_COLUMNS = [
    "item_id",
    "url",
    "label",
    "value_raw",
    "value_display",
    "unit",
    "year",
    "period_start_year",
    "period_end_year",
    "geo",
    "source_name",
    "source_year",
    "sort_order",
    "active",
    "meta_json",
]


class ValidationError(RuntimeError):
    pass


def cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if math.isfinite(value) and value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def parse_bool(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "y"}


def parse_json(value: str, row_number: int) -> dict:
    raw = value.strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValidationError(f"Row {row_number}: invalid meta_json: {exc.msg}") from exc
    if not isinstance(parsed, dict):
        raise ValidationError(f"Row {row_number}: meta_json must be object")
    return parsed


def normalize_row(raw: dict, row_number: int, section: str) -> dict:
    missing = [col for col in REQUIRED_COLUMNS if col not in raw]
    if missing:
        raise ValidationError(f"Row {row_number}: missing columns {missing}")

    item_id = raw["item_id"].strip()
    if not item_id:
        raise ValidationError(f"Row {row_number}: item_id required")

    sort_raw = raw["sort_order"].strip() or "9999"
    try:
        sort_order = int(sort_raw)
    except ValueError as exc:
        raise ValidationError(f"Row {row_number}: sort_order must be integer") from exc

    return {
        "section": section,
        "item_id": item_id,
        "url": raw["url"].strip(),
        "label": raw["label"].strip(),
        "value_raw": raw["value_raw"].strip(),
        "value_display": raw["value_display"].strip(),
        "unit": raw["unit"].strip(),
        "year": raw["year"].strip(),
        "period_start_year": raw["period_start_year"].strip(),
        "period_end_year": raw["period_end_year"].strip(),
        "geo": raw["geo"].strip(),
        "source_name": raw["source_name"].strip(),
        "source_year": raw["source_year"].strip(),
        "sort_order": sort_order,
        "active": parse_bool(raw["active"]),
        "meta_json": parse_json(raw["meta_json"], row_number),
    }


def load_rows_from_worksheet(
    ws, section_name: str, workbook_path: Path
) -> Tuple[List[dict], set[tuple[str, str]]]:
    rows_values = list(ws.iter_rows(values_only=True))
    if not rows_values:
        return [], set()

    header_cells = rows_values[0]
    headers = [cell_to_str(c) for c in header_cells]
    if not any(headers):
        raise ValidationError(
            f"{workbook_path} sheet {section_name!r}: missing header row"
        )

    header_set = {h for h in headers if h}
    missing_cols = [col for col in REQUIRED_COLUMNS if col not in header_set]
    if missing_cols:
        raise ValidationError(
            f"{workbook_path} sheet {section_name!r}: missing columns {missing_cols}"
        )

    rows: List[dict] = []
    seen: set[tuple[str, str]] = set()
    for row_number, values in enumerate(rows_values[1:], start=2):
        raw: Dict[str, str] = {}
        for header, cell in zip_longest(headers, values, fillvalue=None):
            if not header:
                continue
            raw[header] = cell_to_str(cell)

        if not raw.get("item_id", "").strip():
            continue

        row = normalize_row(raw, row_number, section_name)
        key = (row["section"], row["item_id"])
        if key in seen:
            raise ValidationError(
                f"{workbook_path} sheet {section_name!r} row {row_number}: "
                f"duplicate (section,item_id)=({row['section']},{row['item_id']})"
            )
        seen.add(key)
        if row["active"]:
            rows.append(row)
    return rows, seen


def load_rows_from_workbook(workbook_path: Path) -> List[dict]:
    if not workbook_path.is_file():
        raise ValidationError(
            f"Workbook not found: {workbook_path}. "
            f"Expected {CONTENT_WORKBOOK_NAME} in jobs_page/data."
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError(
            "openpyxl required. Install: pip install -r jobs_page/requirements.txt"
        ) from exc

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        all_rows: List[dict] = []
        global_seen: set[tuple[str, str]] = set()
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("_"):
                continue
            section_name = sheet_name.strip()
            if not section_name:
                continue
            ws = wb[sheet_name]
            sheet_rows, keys = load_rows_from_worksheet(ws, section_name, workbook_path)
            duplicates = sorted(keys.intersection(global_seen))
            if duplicates:
                first = duplicates[0]
                raise ValidationError(
                    f"Duplicate (section,item_id) across sheets: ({first[0]},{first[1]})"
                )
            global_seen.update(keys)
            all_rows.extend(sheet_rows)
        return all_rows
    finally:
        wb.close()


def group_by_section(rows: List[dict]) -> Dict[str, List[dict]]:
    grouped: Dict[str, List[dict]] = {}
    for row in rows:
        grouped.setdefault(row["section"], []).append(row)
    for section_rows in grouped.values():
        section_rows.sort(key=lambda item: (item["sort_order"], item["item_id"]))
    return grouped


def format_number_for_unit(value_raw: str, unit: str) -> str:
    if not value_raw:
        return ""
    try:
        if unit == "count":
            return f"{int(float(value_raw)):,}"
        if unit == "currency":
            return f"${int(float(value_raw)):,}"
        if unit == "percent":
            return f"+{float(value_raw):.1f}%"
    except ValueError:
        return value_raw
    return value_raw


def display_value(row: dict) -> str:
    return row["value_display"] or format_number_for_unit(row["value_raw"], row["unit"])


def render_stat_items(rows: List[dict]) -> str:
    blocks = []
    for row in rows:
        css_class = html.escape(row["meta_json"].get("class_name", ""))
        css_attr = f" {css_class}" if css_class else ""
        blocks.append(
            "    <div class=\"stat\">\n"
            f"      <div class=\"stat-num{css_attr}\">{html.escape(display_value(row))}</div>\n"
            f"      <div class=\"stat-label\">{html.escape(row['label'])}</div>\n"
            f"      <div class=\"stat-src\">{html.escape(row['source_name'])} {html.escape(row['source_year'])}</div>\n"
            "    </div>"
        )
    return "\n".join(blocks)


def render_projection_rows(rows: List[dict]) -> str:
    max_growth = 0.0
    parsed_growth = {}
    for row in rows:
        try:
            growth = float(row["value_raw"])
        except ValueError:
            growth = 0.0
        parsed_growth[row["item_id"]] = growth
        if growth > max_growth:
            max_growth = growth
    max_growth = max_growth or 1.0

    blocks = []
    for row in rows:
        growth = parsed_growth[row["item_id"]]
        pct_class = "pct up" if growth >= 0 else "pct"
        growth_text = row["meta_json"].get("growth_display", f"+{growth:.1f}%")
        jobs_text = row["meta_json"].get("jobs_display", row["label"])
        occupation = html.escape(row["label"])
        bar_width = round((growth / max_growth) * 100) if growth > 0 else 0
        is_benchmark = parse_bool(str(row["meta_json"].get("is_benchmark", "false")))
        row_style = ' style="color:var(--text-2)"' if is_benchmark else ""
        fill_class = "bar-fill neutral" if is_benchmark else "bar-fill"
        blocks.append(
            "        <tr>\n"
            f"          <td{row_style}>{occupation}</td>\n"
            f"          <td class=\"mono\"{row_style}>{html.escape(jobs_text)}</td>\n"
            f"          <td class=\"{pct_class}\"{row_style}>{html.escape(growth_text)}</td>\n"
            "          <td>\n"
            "            <div class=\"bar-wrap\">\n"
            "              <div class=\"bar-track\">\n"
            f"                <div class=\"{fill_class}\" style=\"width:{bar_width}%\"></div>\n"
            "              </div>\n"
            "            </div>\n"
            "          </td>\n"
            "        </tr>"
        )
    return "\n".join(blocks)


def render_fl_cells(rows: List[dict], default_num_class: str = "") -> str:
    blocks = []
    for row in rows:
        num_class = row["meta_json"].get("num_class", default_num_class)
        class_suffix = f" {html.escape(num_class)}" if num_class else ""
        blocks.append(
            "      <div class=\"fl-cell\">\n"
            f"        <div class=\"fl-num{class_suffix}\">{html.escape(display_value(row))}</div>\n"
            f"        <div class=\"fl-desc\">{html.escape(row['label'])}</div>\n"
            "      </div>"
        )
    return "\n".join(blocks)


def render_grad_options(rows: List[dict]) -> str:
    return "\n".join(
        f"            <option>{html.escape(row['label'])}</option>" for row in rows
    )


def render_event_items(rows: List[dict]) -> str:
    """Runtime calendar uses Apps Script JSONP (`loadCareerEvents`); workbook rows ignored."""
    return (
        '      <p class="slabel ev-loading-msg" style="margin-bottom:0.75rem;color:var(--text-2);">Loading events…</p>'
    )


def render_footer_text(rows: List[dict]) -> str:
    if not rows:
        return ""
    text = rows[0]["label"]
    return f"    <span class=\"footer-text\">{html.escape(text)}</span>"


def render_source_note(rows: List[dict]) -> str:
    if not rows:
        return ""
    return "      " + html.escape(rows[0]["label"])


def render_context_line(rows: List[dict]) -> str:
    if not rows:
        return ""
    return html.escape(rows[0]["label"])


def render_content_cards(rows: List[dict]) -> str:
    blocks = []
    for row in rows:
        meta = row["meta_json"]
        badge = html.escape(meta.get("badge", ""))
        badge_class = html.escape(meta.get("badge_class", "blue"))
        head = html.escape(meta.get("head", row["label"]))
        body = html.escape(meta.get("body", row["label"]))
        cta = html.escape(meta.get("cta", "Learn more ->"))
        href = html.escape(meta.get("href", "#"), quote=True)
        target = ' target="_blank"' if parse_bool(str(meta.get("target_blank", "false"))) else ""
        badge_html = f"        <div class=\"badge {badge_class}\">{badge}</div>\n" if badge else ""
        blocks.append(
            "      <div class=\"card\">\n"
            f"{badge_html}"
            f"        <div class=\"card-head\">{head}</div>\n"
            f"        <div class=\"card-body\">{body}</div><a class=\"card-cta\" href=\"{href}\"{target}>{cta}</a>\n"
            "      </div>"
        )
    return "\n".join(blocks)


def replace_marker_region(template: str, marker: str, replacement: str) -> str:
    start = f"<!-- GEN:{marker}:start -->"
    end = f"<!-- GEN:{marker}:end -->"
    pattern = re.compile(
        rf"({re.escape(start)})(.*)({re.escape(end)})",
        flags=re.DOTALL,
    )
    match = pattern.search(template)
    if not match:
        raise ValidationError(f"Template missing marker region: {marker}")
    new_body = f"{start}\n{replacement}\n    {end}"
    return template[: match.start()] + new_body + template[match.end() :]


def has_marker_region(template: str, marker: str) -> bool:
    start = f"<!-- GEN:{marker}:start -->"
    end = f"<!-- GEN:{marker}:end -->"
    return start in template and end in template


def render_page(template_html: str, grouped: Dict[str, List[dict]]) -> str:
    rendered = template_html
    renderer_map = {
        "stat_items": render_stat_items,
        "projection_rows": render_projection_rows,
        "prospective_spotlight": lambda rows: render_fl_cells(rows, default_num_class="c-green"),
        "prospective_source_note": render_source_note,
        "current_resources": render_content_cards,
        "jm_grad_options": render_grad_options,
        "current_metro": render_fl_cells,
        "job_match_context_line": render_context_line,
        "event_items": render_event_items,
        "employer_value_cards": render_content_cards,
        "footer_text": render_footer_text,
    }
    for marker, renderer in renderer_map.items():
        if not has_marker_region(rendered, marker):
            continue
        rows = grouped.get(marker, [])
        replacement = renderer(rows)
        rendered = replace_marker_region(rendered, marker, replacement)
    return rendered


def validate_safe_output(output_path: Path, repo_root: Path) -> None:
    resolved_output = output_path.resolve()
    resolved_root = repo_root.resolve()
    if resolved_root not in resolved_output.parents:
        raise ValidationError("Refusing to write outside repository root.")
    if resolved_output.name != "index.html" or resolved_output.parent.name != "jobs_page":
        raise ValidationError("Output path must be jobs_page/index.html for this generator.")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate jobs_page/index.html from jobs_page/data/jobs_page_content.xlsx."
    )
    repo_root = Path(__file__).resolve().parents[2]
    data_dir = repo_root / "jobs_page" / "data"
    default_output = repo_root / "jobs_page" / "index.html"
    parser.add_argument("--template", type=Path, default=default_output)
    parser.add_argument("--output", type=Path, default=default_output)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    try:
        validate_safe_output(args.output, repo_root)
        workbook_path = data_dir / CONTENT_WORKBOOK_NAME
        rows = load_rows_from_workbook(workbook_path)
        grouped = group_by_section(rows)
        template_html = args.template.read_text(encoding="utf-8")
        rendered = render_page(template_html, grouped)
    except (ValidationError, FileNotFoundError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.check:
        current = args.output.read_text(encoding="utf-8")
        if current != rendered:
            print("Generated HTML differs from current output.", file=sys.stderr)
            return 2
        print("Check passed: output is up to date.")
        return 0

    args.output.write_text(rendered, encoding="utf-8")
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
